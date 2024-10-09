import frappe
from frappe.utils import file_lock, now_datetime, get_url
from frappe import _
import requests, json, openpyxl
from six import iteritems, string_types
from frappe.utils import (
	cint
)
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from frappe.utils.xlsxutils import handle_html, ILLEGAL_CHARACTERS_RE
from io import BytesIO

from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

def queue_action(self, action, **kwargs):
	"""Run an action in background. If the action has an inner function,
	like _submit for submit, it will call that instead"""
	# call _submit instead of submit, so you can override submit to call
	# run_delayed based on some action
	# See: Stock Reconciliation
	from frappe.utils.background_jobs import enqueue

	if hasattr(self, '_' + action):
		action = '_' + action

	if file_lock.lock_exists(self.get_signature()):
		frappe.throw(_('This document is currently queued for execution. Please try again'),
			title=_('Document Queued'))
	
	frappe.db.set_value(self.doctype, self.name, 'ais_queue_status', 'Queued',  update_modified=False)
	frappe.db.set_value(self.doctype, self.name, 'ais_queue_failed', 0,  update_modified=False)
	frappe.db.set_value(self.doctype, self.name, 'ais_queueu_comment', '',  update_modified=False)
	frappe.db.set_value(self.doctype, self.name, 'ais_queued_date', now_datetime(),  update_modified=False)
	frappe.db.set_value(self.doctype, self.name, 'ais_queued_by', frappe.session.user,  update_modified=False)
	self.lock()
	enqueue('metactical.custom_scripts.frappe.document.execute_action', doctype=self.doctype, name=self.name,
		action=action, **kwargs)
		
def post_to_rocket_chat(doc, msg, failed=False):
	try:
		rocket_chat_settings = frappe.get_single('Rocket Chat Settings')
		if not rocket_chat_settings.rocket_notification:
			return

		channel_name = rocket_chat_settings.channel_name
		headers = {
			'Content-type': rocket_chat_settings.content_type or 'application/json',
			'X-Auth-Token': rocket_chat_settings.auth_token,
			'X-User-Id': rocket_chat_settings.user_id
		}

		url = "/app/{0}/{1}".format(doc.doctype.lower().replace(" ", "-"), doc.name)
		message = 'A document you submitted has taken too long and has been unquequd. Please resubmit the document and notify the system \
						administrator \n[{0}]({1})'.format(get_url(url), get_url(url))
		
		if failed:
			message = 'A document you submitted has failed. Please see the error in the comment section of the document and fix it \
						\n[{0}]({1})'.format(get_url(url), get_url(url))

		payload = {
			'channel': "#"+channel_name,
			'text': message
		}

		response = requests.post(rocket_chat_settings.url, 
								headers=headers, 
								data=json.dumps(payload))

		if response.status_code == 200:
			pass
		else:
			frappe.log_error(title='Rocket Chat Error', message=response.json())
	except Exception as e:
		frappe.log_error(title='Rocket Chat Error', message=frappe.get_traceback())

def format_json_for_html(data, indent_size=2):
	try:
		# Function to recursively format JSON
		def format_json(data, indent_level):
			lines = []
			for key, value in data.items():
				if isinstance(value, dict):
					# Recursively format nested objects
					lines.append(f'{" " * indent_level * indent_size}"{key}": {{ <br>')
					lines.extend(format_json(value, indent_level + 1))
					lines.append(f'{" " * indent_level * indent_size}}}, <br>')
				elif isinstance(value, list):
					# Handle lists of objects
					lines.append(f'{" " * indent_level * indent_size}"{key}": [ <br>')
					for item in value:
						lines.extend(format_json(item, indent_level + 1))
					lines.append(f'{" " * indent_level * indent_size}] <br>')
				else:
					# Format primitive types (string, number, etc.)
					lines.append(f'{" " * indent_level * indent_size}"{key}": "{value}", <br>')
			return lines
		
		# Start formatting from the top-level object
		formatted_lines = format_json(data, indent_level=1)
		
		# Join all lines with newline characters
		formatted_json = '\n'.join(formatted_lines)
		
		return formatted_json
	
	except json.JSONDecodeError as e:
		return f"Error decoding JSON: {str(e)}"
	except Exception as e:
		return f"Error: {str(e)}"

def create_usaepay_log(doctype, docname, action):
	# Create USAePay Log
	log = frappe.get_doc({
		"doctype": "USAePay Log",
		"date": frappe.utils.now(),
		"reference_docname": docname,
		"action": action,
		"reference_doctype": doctype
	}).insert()

	return log

@frappe.whitelist()
def get_usaepay_account(transaction_key=None, merchant_id=None, lead_source=None):
	usaepay_account = None
	if lead_source:
		usaepay_account = frappe.db.exists("USAePay Accounts", {"lead_source": lead_source})
	elif merchant_id:
		usaepay_account = frappe.db.exists("USAePay Accounts", {"merchant_id": merchant_id})
	elif transaction_key:
		source = frappe.db.get_value("Sales Order", {"neb_usaepay_transaction_key": transaction_key}, "name")
		if source:
			usaepay_account = frappe.db.exists("USAePay Accounts", {"source": source})

	if usaepay_account:
		return frappe.get_doc("USAePay Accounts", usaepay_account)

	return get_default_usaepay_account()

@frappe.whitelist()
def get_default_usaepay_account():
	account = frappe.db.get_value("USAePay Accounts", {"is_default": 1}, "name")
	if not account:
		frappe.throw(_("No default USAePay account found. Please add a default account"))
	else:
		account = frappe.get_doc("USAePay Accounts", account)

	return account

@frappe.whitelist()
def get_customer_payment_information(customer, reference_no, payment_entry):
	from metactical.custom_scripts.usaepay.usaepay_api import get_token_hash

	# get existing credit card tokens
	tokens = []
	lead_source = ""
	references = frappe.get_doc("Payment Entry", payment_entry).references
	for ref in references:
		if ref.reference_doctype == "Sales Order" and ref.reference_name:
			lead_source = frappe.db.get_value("Sales Order", ref.reference_name, "source")
			break

	if frappe.db.exists("Customer CC", customer):
		tokens = frappe.get_list("Customer CC Tokens", {"parent": customer}, ["name", "label", "token", "cc_number"])

	usaepay_settings = get_usaepay_account(reference_no, None, lead_source)
	payment_form_url = usaepay_settings.payment_form_url
	billing_address = get_customer_address(customer)

	form_hash = get_token_hash(usaepay_settings, "1234")
	# form_hash = form_hash[6:] if form_hash else None
	
	if not form_hash:
		frappe.log_error(title="Metactical Settings Error", message="Failed to generate form hash. Please add usaepay key and secret")
		frappe.throw(_("Failed to generate form hash. Please check the MetaTactical settings"))

	frappe.response["tokens"] = tokens
	frappe.response["payment_form_url"] = payment_form_url
	frappe.response["address"] = billing_address
	frappe.response["hash"] = form_hash

def get_customer_address(customer):
	addresses = frappe.db.sql("""SELECT
			address_line1, address_line2, city, state, 
			country, phone, company, pincode, 
			phone, address_type, 
			is_shipping_address, is_primary_address
		FROM
			`tabAddress`
		JOIN
			`tabDynamic Link` ON `tabDynamic Link`.parent = `tabAddress`.name
		WHERE
			`tabDynamic Link`.link_doctype = 'Customer' AND
			`tabDynamic Link`.link_name = %(customer)s 
		""", {"customer": customer}, as_dict=1)

	grouped_address = {}
	billing_address = None
	shipping_address = None
	for address in addresses:
		if billing_address and shipping_address:
			break

		if address.get("is_primary_address"):
			billing_address = address
		elif address.get("is_shipping_address"):
			shipping_address = address

		if address.address_type not in grouped_address:
			grouped_address[address.address_type] = []

		grouped_address[address.address_type].append(address)
		
	if not billing_address:
		billing_address = grouped_address["Billing"][0] if "Billing" in grouped_address else None

	if not shipping_address:
		shipping_address = grouped_address["Shipping"][0] if "Shipping" in grouped_address else None
	
	# add customer personal information to the address
	customer_info = frappe.db.get_value("Customer", customer, ["ais_company", "first_name", "last_name"], as_dict=1)
	if customer_info:
		billing_address.update(customer_info) if billing_address else None
		shipping_address.update(customer_info) if shipping_address else None

	return {
		"billing": billing_address,
		"shipping": shipping_address
	}

@frappe.whitelist()
def export_query(data, sub_headers=[]):
	from frappe.desk.query_report import run, get_columns_dict, handle_duration_fieldtype_values
	report_name = data.get("report_name")
	filters = data.get("filters")

	frappe.permissions.can_export(
		frappe.get_cached_value("Report", report_name, "ref_doctype"),
		raise_exception=True,
	)

	file_format_type = data.get("file_format_type")
	custom_columns = frappe.parse_json(data.get("custom_columns", "[]"))
	include_indentation = data.get("include_indentation")

	if file_format_type == "Excel":
		data = run(report_name, filters, custom_columns=custom_columns)
		data = frappe._dict(data)
		if not data.columns:
			frappe.respond_as_web_page(
				_("No data to export"),
				_("You can try changing the filters of your report."),
			)
			return

		columns = get_columns_dict(data.columns)
		data["result"] = handle_duration_fieldtype_values(data.get("result"), data.get("columns"))

		xlsx_data, column_widths = build_xlsx_data(columns, data, filters.get("date"))
		xlsx_file = make_xlsx(xlsx_data, sub_headers, "Query Report", column_widths=column_widths)

		# get current date as string format
		current_date = now_datetime().strftime('%d-%m-%Y')

		frappe.response["filename"] =  current_date + " " +report_name  + ".xlsx"
		frappe.response["filecontent"] = xlsx_file.getvalue()
		frappe.response["type"] = "binary"

def build_xlsx_data(columns, data, date):
	result =[]
	column_widths = []
	columns = []
	for column in data.columns:
		if column.get("hidden"):
			continue
		columns.append(column.get("label"))
		column_width = cint(column.get("width", 0))
		# to convert into scale accepted by openpyxl
		column_width /= 10
		column_widths.append(column_width)
	
	result.append(["Date", frappe.utils.get_datetime(date).strftime('%d-%b-%Y')])
	result.append(columns)

	# build table from result
	for row_idx, row in enumerate(data.result):
		# only pick up rows that are visible in the report
		row_data = []
		if isinstance(row, dict):
			for col_idx, column in enumerate(data.columns):
				if column.get("hidden"):
					continue
				label = column.get("label")
				fieldname = column.get("fieldname")
				cell_value = row.get(fieldname, row.get(label, ""))
				row_data.append(cell_value)
		elif row:
			row_data = row

		result.append(row_data)

	return result, column_widths

# return xlsx file object
def make_xlsx(data, sub_headers, sheet_name, wb=None, column_widths=None):
	from openpyxl.styles import Alignment
	from openpyxl.styles import PatternFill

	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)
	headers = []

	for i, row in enumerate(data):
		clean_row = []
		for col, item in enumerate(row):
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)

			clean_row.append(value)
			if (item in sub_headers) or i == 0:
				headers.append(i+1)

		ws.append(clean_row)
		ws.row_dimensions[i+1].height = 20

	for header in headers:
		if (header == 1):
			ws.merge_cells(start_row=header, start_column=2, end_row=header, end_column=len(row))
		else:
			ws.merge_cells(start_row=header, start_column=1, end_row=header, end_column=len(row))
		
		for rows in ws.iter_rows(min_row=header, max_row=header, min_col=1, max_col=len(row)):
			for cell in rows:
				cell.alignment = Alignment(wrap_text=True)
				cell.alignment = Alignment(vertical="center", horizontal="center")
				cell.fill = PatternFill(start_color="ffd966", end_color="ffd966", fill_type="solid")

	set_border(len(row) + 1, len(data), ws)
	set_auto_width(ws)

	# set font
	for row in ws.iter_rows(min_row=1, max_row=len(data), min_col=1, max_col=len(data[1])):
		if row[0].row == 2:
			for cell in row:
				cell.font = Font(size=10, bold=True, name="Calibri")
		elif row[0].row in headers:
			for cell in row:
				cell.font = Font(size=11, name="Calibri")
		else:
			for cell in row:
				cell.font = Font(size=10, name="Segoe UI")

	# Auto fit rows to their content
	for row in ws.iter_rows(min_row=1, max_row=len(data), min_col=1, max_col=len(data[1])):
		ws.row_dimensions[row[0].row].auto_size = True

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def set_auto_width(ws):
	for column_cells in ws.columns:
		max_length = 0
		col_letter = get_column_letter(column_cells[0].column)  # Get the column letter
		
		for cell in column_cells:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
			
		adjusted_width = max_length if max_length > 5 else 6
		ws.column_dimensions[col_letter].width = adjusted_width

def set_border(columns, rows, ws):
	thin_border = Border(
		left=Side(style='thin', color='f2f2f2'),
		right=Side(style='thin', color='f2f2f2'),
		top=Side(style='thin', color='f2f2f2'),
		bottom=Side(style='thin', color='f2f2f2')
	)

	for row in range(1, rows):
		for col in range(1, columns):
			cell = ws.cell(row=row, column=col)
			cell.border = thin_border

# check if all sales invoices are paid for a sales order
def check_si_payment_status_for_so(sales_order):
	all_invoices_paid = False
	
	invoices = frappe.db.sql("""SELECT
								invoice.name AS invoice_name, invoice.status, invoice.grand_total
							FROM
								`tabSales Invoice Item`  AS item
							LEFT JOIN
								`tabSales Invoice` AS invoice ON invoice.name = item.parent
						  	WHERE
						  		item.sales_order = %(sales_order)s
						  GROUP BY invoice.name, invoice.status, invoice.grand_total""", 
				{"sales_order": sales_order}, as_dict=True)
	
	order_grand_total = frappe.db.get_value("Sales Order", sales_order, "grand_total")
	invoices_total = 0

	if invoices and len(invoices) > 0:
		for invoice in invoices:
			if invoice.status != "Paid":	
				all_invoices_paid = False
				break
			else:
				invoices_total += invoice.grand_total

	if not all_invoices_paid and invoices_total == order_grand_total:
		all_invoices_paid = True

	return all_invoices_paid

def get_customer_email_and_phone(customer):
    contacts = frappe.db.sql("""select c.email_id, phone, c.mobile_no
								from `tabContact` c
								INNER JOIN `tabDynamic Link` dl on dl.parent=c.name
								INNER Join `tabCustomer` cs on dl.link_name=cs.name
								where  dl.link_doctype="Customer" and cs.name = "{0}"
                                ORDER BY c.creation desc
                                """.format(customer), as_dict=True)
            

    if len(contacts):
        return contacts
    else:
        return None

def search_customer_by_phone_email(phone_number, email):
    email_filter = ""
    if email:
        email_filter = f"AND c.email_id like '%{email}%'"
    
    phone_filter = ""
    if phone_number:
        phone_filter = f"AND (c.phone like '%{phone_number}%' or c.mobile_no like '%{phone_number}%')"

    customers = frappe.db.sql(f"""select cs.name
                                from `tabContact` c
                                INNER JOIN `tabDynamic Link` dl on dl.parent=c.name
                                INNER Join `tabCustomer` cs on dl.link_name=cs.name
                                where  dl.link_doctype="Customer" {email_filter} {phone_filter}
                                """, as_dict=True)

    if len(customers):
        return [customer.get('name') for customer in customers]
    else:
        return None