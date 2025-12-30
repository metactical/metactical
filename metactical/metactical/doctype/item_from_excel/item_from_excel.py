# Copyright (c) 2024, Techlift Technologies and contributors
# For license information, please see license.txt

import frappe
import os
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xls_file_from_attached_file
from metactical.custom_scripts.utils.metactical_utils import queue_action
from openpyxl import load_workbook
from io import BytesIO
from frappe import _, msgprint
from frappe.model.docstatus import DocStatus
import requests

# from erpnext.controllers.item_variant import (
# 	make_variant_item_code
# )

class ItemFromExcel(Document):
	def save(self):
		if self.docstatus == DocStatus.submitted() and \
			self.ais_queue_status and self.ais_queue_status != "Queued":
			msgprint(
				_(
					"The task has been enqueued as a background job. In case there is \
					any issue on processing in background, the system will add a comment \
					about the error on this document and revert to the Draft stage"
				)
			)
			queue_action(self, "submit", timeout=2000)
		else:
			super().save()

	def validate(self):
		file_content = self.check_file()
  
		if not self.all_valid and self.docstatus == DocStatus.submitted():
			frappe.throw("Please correct the errors in the file before submitting.")

		linked_doctypes, item_field_map, required_fields = get_doctype_information()

		# check if all required fields are present
		if len(file_content) < 3:
			frappe.throw(f"Required number of sheets not found in the uploaded file. Expected 3 (Template, Variant, Spec labels), found {len(file_content)}")

		for field in required_fields:
			if field not in file_content[0][0]:
				frappe.throw(f"Required field {field} not found in the uploaded file")

			if field not in file_content[1][0]:
				frappe.throw(f"Required field {field} not found in the uploaded file")
		
		self.check_mandatory_fields(file_content[0], required_fields, is_template=True)
		self.check_specs_availability(file_content[2])
		self.check_mandatory_fields(file_content[1], required_fields, is_template=False)
  
	def check_specs_availability(self, data):
		headers = data[0]
		specs = headers[1:]  # assuming first column is something else like Item Code
		missing_specs = []
  
		for spec in specs:
			if not frappe.db.exists("Website Specification Label", spec):
				missing_specs.append(spec)
    
		if missing_specs:	
			frappe.throw(f"The following Specification Labels are missing in the ERP: <b>{', '.join(missing_specs)}</b>")


	def check_mandatory_fields(self, data, mandatory_fields, is_template):
		headers = data[0]
		mandatory_fields_index = [i for i, x in enumerate(headers) if x in mandatory_fields]

		for d in data[1:]:
			for i in mandatory_fields_index:
				if not d[i] and d[0]:
					frappe.throw(f"Value missing for field {headers[i]} at row {data.index(d) + 1}")

	def create_item(self, file_content, item_field_map, linked_dcts, is_template, templates_with_ai_request={}):
		self.attributes = ""
		self.attribute_values = ""

		data = file_content[0] if is_template else file_content[1]
		specs = file_content[2]
		spec_labels = self.prepare_website_specifications_from_data(specs)
  
		# collect all the last variants of the template and trigger AI update after the price list is created
		template_with_last_variant = {}
		variant_of_header = data[0].index("Variant Of") if "Variant Of" in data[0] else -1
  
		for d in data[1:] if not is_template else []:
			if d and d[0]:
				template_with_last_variant[d[variant_of_header]] = d[0]
    
		# Helper function to initialize data structures for a new item
		def initialize_item_data():
			return frappe.new_doc("Item"), "", {ld: [] for ld in linked_dcts}, {ld: {} for ld in linked_dcts}, [price_list_headers]

		# Helper function to update the item's fields based on the provided field and value
		def update_item_field(item, field, value):
			if field in item_field_map and value is not None:
				item.set(item_field_map[field], value)
    
		def validate_mandatory_specs(item, spec_labels):
			item_group = frappe.get_cached_doc("Item Group", item.item_group)
			website_specifications = item_group.get("neb_website_specifications") if item_group else []
   
			mandatory_specs = [spec.label for spec in website_specifications if spec.mandatory]
			item_codes_with_spec = spec_labels.keys()
   
			# check if there are mandatory specs for the item's item group
			if mandatory_specs and item.item_code in item_codes_with_spec:
				item_specs = spec_labels[item.item_code]
				for ms in mandatory_specs:
					if ms not in item_specs:
						frappe.throw(f"Mandatory specification <b>'{ms}'</b> missing for item <b>'{item.item_code}'</b>")				

		# Helper function to process linked doctypes and update temporary child table values
		def process_linked_doctypes(fields, row):
			for doctype in updated_linked_doctypes_to_map:
				parent_label = get_parent_label(linked_dcts, doctype)

				for i, field in enumerate(fields):
					if field == "Attribute (Variant Attributes)" and row[i]:
						self.attributes = row[i]
					elif field == "Attribute Value (Variant Attributes)" and row[i]:
						self.attribute_values = row[i]
					else:
						if field and field.endswith(f"({parent_label})") and row[i] is not None:
							child_table = get_key_from_value(linked_dcts, doctype)
							child_table_field = updated_linked_doctypes_to_map[doctype][field]
							temp_child_table_values[child_table][child_table_field] = row[i]
       
			# check if the number of attributes and attribute values match
			if self.attributes and self.attribute_values and not is_template:
				attributes = self.attributes.split(', ') if "," in self.attributes else [self.attributes]
				attribute_values = self.attribute_values.split(', ') if not type(self.attribute_values) == int else [self.attribute_values]

				if len(attributes) != len(attribute_values):
					frappe.throw(f"Number of attributes and attribute values do not match for row {data.index(row) + 1}")
		  
		# Extract field names and price list headers from the first row
		fields = data[0]
		price_list_index = fields.index("Price List") if "Price List" in fields else -1
		price_list_headers, cost_column_index, item_name_column = self.get_price_list_headers(fields, is_template)
		linked_doctypes_to_map, updated_linked_doctypes_to_map = get_linked_doctypes(linked_dcts, fields)

		# Initialize item and related data structures
		item, item_code, child_table_values, temp_child_table_values, price_list_rows = initialize_item_data()

		# Iterate over each row in the data (excluding the first row)
		for index, row in enumerate(data[1:]):
			# Determine the row to check depending on whether the item name or item code
			row_to_check = row[0] if is_template else row[item_name_column]

			# If starting a new item, save the current one and reinitialize variables
			if index > 0 and row_to_check and item_code != row_to_check:
				is_last_item_of_template = True if item.variant_of and template_with_last_variant.get(item.variant_of) == item.item_code else False
				validate_mandatory_specs(item, spec_labels)
				self.save_item(item, 
                   				child_table_values, 
                   				is_template, price_list_rows, 
                       			self.attributes, 
                          		self.attribute_values, 
                            	data[index][price_list_index], 
                             	is_last_item_of_template,
                              	templates_with_ai_request,
								spec_labels
							  )	
    
				item, item_code, child_table_values, temp_child_table_values, price_list_rows = initialize_item_data()

			prices = []
			# Iterate over each field in the current row
			for i, field in enumerate(fields):
				# Process item fields before the cost column index
				if i < cost_column_index or cost_column_index == -1:
					if index == 0:
						item_code = row_to_check
					update_item_field(item, field, row[i])
	
				# Process price fields for non-template items
				elif not is_template and i >= cost_column_index:
					prices.append(row[i])

			# Append prices to the price list if there are any valid prices
			if prices and not all(p is None for p in prices):
				price_list_rows.append(prices)
			
			# Process linked doctypes and update child table values
			process_linked_doctypes(fields, row)
			
			# Transfer temporary child table values to the main child table values
			for child_table, temp_values in temp_child_table_values.items():
				if temp_values:
					child_table_values[child_table].append(temp_values.copy())
					temp_child_table_values[child_table] = {}

		if item:
			is_last_item_of_template = True if item.variant_of and template_with_last_variant.get(item.variant_of) == item.item_code else False
			validate_mandatory_specs(item, spec_labels)
			# Save the last item after the loop
			self.save_item(item, 
                  			child_table_values, 
                  			is_template, 
                     		price_list_rows, 
                       		self.attributes, 
                         	self.attribute_values, 
                          	data[-1][price_list_index],
							is_last_item_of_template,
							templates_with_ai_request,
							spec_labels
						)
   
	def prepare_website_specifications_from_data(self, sheet_data):
		if not sheet_data or len(sheet_data) < 2:
			return {}
		
		# First row is headers
		headers = sheet_data[0]
		
		if not headers or len(headers) < 2:
			return {}
		
		# First column should be Item Group, rest are specification labels
		spec_labels = []
		for i, header in enumerate(headers[1:], start=1):
			if header and str(header).strip():
				spec_labels.append({
					"index": i,
					"label": str(header).strip()
				})
		
		if not spec_labels:
			return {}
		
		# Extract specifications for each item
		item_specifications = {}
		
		for row in sheet_data[1:]:  # Skip header row
			if not row or not row[0]:  # Skip empty rows
				continue
			
			item_code = str(row[0]).strip()
			specifications = []
			
			# Check each specification label column
			for spec_info in spec_labels:
				col_idx = spec_info["index"]
				
				if len(row) > col_idx and row[col_idx] is not None:
					description = str(row[col_idx]).strip()
     
					if not frappe.db.exists("Website Spec Label Descriptions", {"parent": spec_info["label"], "description": description}):
						frappe.throw(f"Specification description '{description}' for label '{spec_info['label']}' does not exist in ERP.")
					
					if description:  # Only add if description is not empty
						specifications.append({
							"label": spec_info["label"],
							"description": description
						})
			
			# Only add to result if item has specifications
			if specifications:
				item_specifications[item_code] = specifications
		
		frappe.logger().info(f"Prepared specifications for {len(item_specifications)} items")
		return item_specifications


	def save_item(self, item, 
               			child_table_values, 
                  		is_template, 
                    	price_list_rows, 
                     	attributes, 
                      	attribute_values, 
                       	price_list, 
                        is_last_item_of_template, 
                        templates_with_ai_request,
						spec_labels
					  ):
     
		child_table_values = remove_duplicate_child_table_values(child_table_values)
		item = add_child_table_values_to_item(item, child_table_values, is_template, attributes, attribute_values)
		item = self.add_website_specifications_to_item(item, spec_labels)

		# generate the item code if it is a variant
		# if not (item.item_code and is_template):
		# 	template_item_name = frappe.db.get_value("Item", item.variant_of, "item_name")
		# 	make_variant_item_code(item.variant_of, template_item_name, item)

		# check if the template item already exists. if it does, skip creating the template item
		if is_template and frappe.db.exists("Item", item.item_code):
			return

		# set the retail sku suffix from the item code
		# item.ifw_retailskusuffix = item.item_code
  
		frappe.flags.in_import = True
		frappe.flags.item_from_excel = True
		item.insert()
		supplier = item.supplier_items[0].supplier if item.supplier_items else None
		if supplier:
			self.create_item_defaults(item, supplier)
		
		# add the item_code, retail sku, and supplier to the price list rows
		price_list_rows = self.add_item_details_to_price_list(price_list_rows, item, price_list)
		
		if not is_template:
			self.create_item_price(price_list_rows)
   
		# trigger AI update for the last variant of the template
		if is_last_item_of_template and item.variant_of in templates_with_ai_request:
			template_item = frappe.get_doc("Item", item.variant_of)
			template_item.request_ai_suggestion = 1
			frappe.flags.in_import = False
			template_item.save()
   
		frappe.flags.item_from_excel = False
   
	def add_website_specifications_to_item(self, item, item_specifications):
		"""
		Add website specifications to item from the prepared data
		
		Args:
			item: Item document
			item_specifications: Dict with item specifications
		"""
		if not item_specifications or item.item_code not in item_specifications:
			return item
		
		specs = item_specifications[item.item_code]
		all_specs = []
		if item.get("item_group"):
			all_specs = frappe.get_all("MT Item Website Specification", 
									   filters={"parent": item.item_group}, 
									   fields=["label", "description", "mandatory"])
   
   
  
		# Add specifications from dat
		specs_from_sheet = []
		for spec in specs:
			# Add to child table
			specs_from_sheet.append(spec["label"])
			item.append("neb_website_specifications", {
				"label": spec["label"],
				"description": spec["description"],
				"mandatory": 0  # Default to not mandatory
			})

   
		for spec in all_specs:
			if spec.label not in specs_from_sheet:
				item.append("neb_website_specifications", {
					"label": spec.label,
					"description": spec.description
				})
    
		return item
   
	def create_item_defaults(self, item, supplier):
		item_default_name = frappe.db.exists("Item Default", {"parent": item.name, "company": frappe.db.get_default("company")})		
		if not item_default_name:
			frappe.get_doc({
				"doctype": "Item Default",
				"parent": item.name,
				"parenttype": "Item",
				"parentfield": "item_defaults",
				"default_supplier": supplier,
				"company": frappe.db.get_default("company")
			}).insert()
		else:
			frappe.db.set_value("Item Default", item_default_name, "default_supplier", supplier)

	def add_item_details_to_price_list(self, price_list_rows, item, price_list):
  
		for i, plr in enumerate(price_list_rows[1:]):
			plr.insert(0, price_list)
			if item.supplier_items:
				plr.insert(0, item.supplier_items[0].supplier)
			else:
				plr.insert(0, "")

			plr.insert(0, item.ifw_retailskusuffix)
			plr.insert(0, item.item_code)
			plr.insert(0, item.item_code)
		
		return price_list_rows

	def get_price_list_headers(self, headers, is_template):
		price_list_headers = ["Item Code", "ERPSKU", "Retail Sku", "Supplier", "Price List"]
		cost_column_index = -1
		item_name_column = -1

		if not is_template:
			for i, header in enumerate(headers):
				if header == "Item Name":
					item_name_column = i
				elif header == "Cost":
					cost_column_index = i
				
				if cost_column_index != -1:
					price_list_headers.append(header)

			if cost_column_index == -1:
				frappe.throw("Cost column not found in variant sheet")

		return price_list_headers, cost_column_index, item_name_column


	def create_item_price(self, data):
		# add the items to the price list

		headers = data[0]
		price_lists = []
		price_list_headers_index = []
  
		data, headers = self.update_data_with_supplier_price_lists(data, headers)
		for i, header in enumerate(headers):
			if header not in ["Item Code", "Item Name", "Item Group", "TemplateSKU", "ERPSKU", "Supplier", "Price List"]:
				price_list = frappe.db.exists("Price List", {"name": header})

				if price_list:
					price_lists.append(header)
					price_list_headers_index.append(i)

		# get all the default price lists for the suppliers

		# get price lists if all the cells in the column are empty
		columns_to_remove = []
		for plhi in price_list_headers_index:
			found = False
			for data_row in data[1:]:
				if data_row[plhi] is not None:
					found = True
					break

			if not found:
				columns_to_remove.append(plhi)

		if price_lists == [] or len(price_lists) == len(columns_to_remove):
			return

		# remove all the price list columns that have all empty/None cells
		updated_data = []
		for row in data:
			updated_data.append([row[i] for i in range(len(row)) if i not in columns_to_remove])
		
		from metactical.metactical.doctype.item_price_from_excel.item_price_from_excel import ItemPriceFromExcel
		ItemPriceFromExcel.create_price_entries(self, updated_data, True)

	# get supplier's price list and append it at the end of the headers and update the data with the cost
	def update_data_with_supplier_price_lists(self, data, headers):
		headers = headers.copy()
		price_list_index = headers.index("Price List")
		cost_column_index = headers.index("Cost")
  
		# get all the suppliers from the excel (Price List sheet)
		price_lists = [row[price_list_index] for row in data[1:] if row[price_list_index] is not None]
		price_lists = list(set(price_lists))
		
		# if there are suppliers with price lists, update the data with the price list and the cost
		if len(price_lists) > 0:
			for i, d in enumerate(data):
				# add empty values to all the rows to match the price list columns because the suppier price lists are added in the header
				if price_lists and i != 0:
					d += [None] * len(price_lists)

				if i == 0:
					for pl in price_lists:
						if pl not in headers:
							headers.append(pl)

				else:
					cost = d[cost_column_index]
					index = headers.index(d[price_list_index])
					d[index] = cost

			data[0] = headers
   
		return data, headers

	def on_submit(self):
		file_content = self.check_file()
		linked_doctypes, item_field_map, required_fields = get_doctype_information()

		try:
			# create template items
			self.create_item(file_content, item_field_map, linked_doctypes, True)
			
			template_headers = file_content[0][0]
			request_ai_suggestion_index = template_headers.index("Request AI Suggestion For Slugs and Descriptions") if "Request AI Suggestion For Slugs and Descriptions" in template_headers else -1
			templates_with_ai_request = {}
   
			if request_ai_suggestion_index != -1:
				for d in file_content[0][1:]:
					if d and d[0] and d[request_ai_suggestion_index]:
						templates_with_ai_request[d[0]] = True

			# create variant items
			self.create_item(file_content, item_field_map, linked_doctypes, False, templates_with_ai_request)
		
			frappe.db.commit()
		except Exception as e:
			frappe.db.rollback()
			frappe.clear_last_message()
			frappe.log_error(title="Error creating items", message=frappe.get_traceback())
			frappe.throw(f"Error creating items: {e}")
			self.db_set("ais_queueu_comment", e)

	def read_file(self):
		file_path = self.excel_file
		extn = os.path.splitext(file_path)[1][1:]

		file_content = None

		file_name = frappe.db.get_value("File", {"file_url": file_path})
		if file_name:
			file = frappe.get_doc("File", file_name)
			file_content = file.get_content()
		
		return file_content, extn

	def check_file(self):
		file_content, extn = self.read_file()
		if extn == "xlsx":
			file_content = self.read_xlsx_file_from_attached_file(fcontent=file_content)
		elif extn == "xls":
			file_content = read_xls_file_from_attached_file(file_content)
		else:
			frappe.throw("Only xls and xlsx files are supported.")
		
		# Cleaned data
		cleaned_data = remove_all_none_rows(file_content)
		return cleaned_data

	def read_xlsx_file_from_attached_file(self, fcontent=None, filepath=None):
		if fcontent:
			filename = BytesIO(fcontent)
		elif filepath:
			filename = filepath
		else:
			return

		sheets = []
		wb1 = load_workbook(filename=filename, read_only=True, data_only=True)
		for ws in wb1.worksheets:
			rows = []
			for row in ws.iter_rows():
				tmp_list = []
				for cell in row:
					tmp_list.append(cell.value)
				rows.append(tmp_list)
			sheets.append(rows)

		return sheets

def get_linked_doctypes(linked_dcts, fields):
	linked_doctypes_to_map = []

	# get all the fields that are used in the excel
	for field, prop in linked_dcts.items():
		for field in fields:
			if field:
				if field.endswith("("+prop['label']+")"):
					linked_doctypes_to_map.append(prop['doctype'])

	linked_doctypes_to_map = list(set(linked_doctypes_to_map))
	linked_doctypes_map = get_linked_doctypes_map(linked_doctypes_to_map)

	# prepare linked doctype fields to match the columns in the excel
	# eg. instead of Attribute (Item Variant Attribute) it should be Attribute (Attributes)
	updated_linked_doctypes_to_map = {}
	for doctype in linked_doctypes_to_map:
		for field in linked_doctypes_map[doctype]:
			if field:
				if not doctype in updated_linked_doctypes_to_map:
					updated_linked_doctypes_to_map[doctype] = {}

				parent_label = get_parent_label(linked_dcts, doctype)
				
				if not field["label"]:
					continue

				label = field["label"] + " ("+parent_label+")"
				updated_linked_doctypes_to_map[doctype][label] = field["fieldname"]

	return linked_doctypes_to_map, updated_linked_doctypes_to_map

def add_child_table_values_to_item(item, child_table_values, is_template, attributes, attribute_values):
	from frappe.client import validate_link
	child_table_values['attributes'] = [] 	
	add_attributes_to_child_table(item, child_table_values, attributes, attribute_values, is_template)
	
	for child in child_table_values:
		if child == "attributes" and is_template:
			for attr in child_table_values[child]:
				is_numeric = frappe.db.get_value("Item Attribute", {"name": attr["attribute"]}, "numeric_values")
				if is_numeric:
					props = frappe.db.get_value("Item Attribute", {"name": attr["attribute"]}, ["from_range", "to_range", "increment"], as_dict=1)
					attr["from_range"] = props["from_range"]
					attr["to_range"] = props["to_range"]
					attr["increment"] = props["increment"]
					attr["numeric_values"] = 1
		if child == "item_detail" and is_template and child_table_values[child]:
			price_lists = child_table_values[child][0].get("price_list")
			if price_lists:
				price_lists = price_lists.split(',') if len(price_lists) > 1 else []
				child_table_values[child] = []
				for pl in price_lists:
					if validate_link("Price List", pl):
						child_table_values[child].append({
							"price_list": pl.strip()
						})
     
		if len(child_table_values[child]):
			item.set(child, child_table_values[child])
			child_table_values[child] = []

	return item

def add_attributes_to_child_table(item, child_table_values, attributes, attribute_values, is_template=False):
	attr_list = attributes.split(', ') if "," in attributes else [attributes]
	attr_values_list = attribute_values.split(', ') if not type(attribute_values) == int else [attribute_values]

	if is_template:
		for attr in attr_list:
			is_numeric = frappe.db.get_value("Item Attribute", {"name": attr.strip()}, "numeric_values")
			child_table_values['attributes'].append({
				"attribute": attr.strip(),
				"attribute_value": None,
				"numeric_values": is_numeric
			})
	else:
		for i, attr in enumerate(attr_list):
			is_numeric = frappe.db.get_value("Item Attribute", {"name": attr.strip()}, "numeric_values")
			child_table_values['attributes'].append({
				"variant_of": item.variant_of,
				"attribute": attr.strip(),
				"attribute_value": attr_values_list[i].strip() if type(attr_values_list[i]) == str else attr_values_list[i],
				"numeric_values": is_numeric
			})

def get_doctype_information():
	item_doctype_meta = frappe.get_meta("Item")
	item_field_map = {}
	linked_doctypes = {}
	required_fields = []

	# create matching dict for item doctype {label: fieldname} 
	for field in item_doctype_meta.fields:
		item_field_map[field.label] = field.fieldname

		if field.fieldtype == "Table":
			linked_doctypes[field.fieldname] = {
				"doctype": field.options,
				"label": field.label
			}

			if field.reqd:
				required_fields.append(field.label)

	return linked_doctypes, item_field_map, required_fields

def get_key_from_value(d, value):
	for key, val in d.items():
		if val["doctype"] == value:
			return key
	return None

def get_linked_doctypes_map(linked_doctypes):
	linked_doctypes_map = {}
	for doctype in linked_doctypes:
		meta = frappe.get_meta(doctype)
		field_map = {}
		for field in meta.fields:
			field_map["fieldname"] = field.fieldname
			field_map["label"] = field.label

			if doctype not in linked_doctypes_map:
				linked_doctypes_map[doctype] = []

			linked_doctypes_map[doctype].append(field_map)
			field_map = {}

	return linked_doctypes_map

def get_parent_label(linked_doctypes, doctype):
	for d in linked_doctypes:
		if linked_doctypes[d]["doctype"] == doctype:
			return linked_doctypes[d]["label"]
	return None

# Function to remove rows with all None values
def remove_all_none_rows(data):
	return [[row for row in table if not all(value is None for value in row)] for table in data]

def remove_duplicate_child_table_values(data):
	return {key: [dict(t) for t in {tuple(d.items()) for d in data[key]}] for key in data}





# Refactored extract_and_validate_excel_data - Multiple Helper Methods
def get_workbook_from_file(file_url):
	"""Get and load Excel workbook from file URL"""
	file_name = frappe.db.get_value("File", {"file_url": file_url})
	if not file_name:
		frappe.throw("File not found")
	
	file_doc = frappe.get_doc("File", file_name)
	file_content = file_doc.get_content()
	
	wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
	return wb


def validate_required_sheets(wb):
	"""Check if required sheets exist in workbook"""
	required_sheets = ['Template', 'Variants']
	for sheet in required_sheets:
		if sheet not in wb.sheetnames:
			frappe.throw(f"Required sheet '{sheet}' not found in the Excel file")


def get_sheet_headers(sheet):
	"""Extract headers from first row of sheet"""
	for idx, row in enumerate(sheet.iter_rows(values_only=True)):
		if idx == 0:
			return list(row)
	return []


def get_price_list_columns(headers):
	"""
	Get price list columns (all columns after 'Cost' column)
	
	Returns:
		dict: {price_list_name: column_index}
		int: cost_column_index
	"""
	cost_column_index = -1
	if 'Cost' in headers:
		cost_column_index = headers.index('Cost')
	else:
		frappe.throw("Cost column not found in Variants sheet")
	
	price_list_columns = {}
	for i, header in enumerate(headers):
		if i > cost_column_index and header and header.strip():
			price_list_columns[header.strip()] = i
	
	if not price_list_columns:
		frappe.throw("No price list columns found after Cost column in Variants sheet")
	
	return price_list_columns, cost_column_index


def get_validation_configs():
	"""
	Get enabled validation configurations
	
	Returns:
		dict: {price_list_name: config}
	"""
	validation_configs = frappe.get_all(
		"Item Import Validation",
		filters={"enabled": 1},
		fields=["name", "api_url", "api_key", "price_list"]
	)
	
	if not validation_configs:
		frappe.throw("No enabled Item Import Validation configurations found")
	
	# Create mapping
	price_list_to_config = {}
	for config in validation_configs:
		price_list_to_config[config.price_list] = config
	
	return price_list_to_config


def get_column_indices(headers):
	"""
	Get column indices for important fields in Variants sheet
	
	Returns:
		dict: {field_name: column_index}
	"""
	return {
		'item_group': headers.index('Item Group') if 'Item Group' in headers else -1,
		'brand': headers.index('Brand') if 'Brand' in headers else -1,
		'availability_rule': headers.index('VariantAvailabilityRule') if 'VariantAvailabilityRule' in headers else -1
	}


def extract_specs_from_third_sheet(wb):
	"""
	Extract specification labels from third sheet
	
	Third sheet structure:
	- Column 1: Item Group
	- Columns 2+: Specification Labels
	
	Returns:
		dict: {item_group: set of spec labels}
	"""
	specs_by_item_group = {}
	
	if len(wb.sheetnames) < 3:
		frappe.logger().info("No third sheet found for specifications")
		return specs_by_item_group
	
	specs_sheet = wb[wb.sheetnames[2]]
	spec_headers = get_sheet_headers(specs_sheet)
	
	if not spec_headers or len(spec_headers) < 2:
		frappe.logger().warning("Third sheet has invalid structure")
		return specs_by_item_group
	
	# Get spec labels (columns 2+)
	spec_labels = []
	for i, header in enumerate(spec_headers[1:], start=1):
		if header and str(header).strip():
			spec_labels.append({
				"index": i,
				"label": str(header).strip()
			})
	
	if not spec_labels:
		frappe.logger().warning("No specification labels found")
		return specs_by_item_group
	
	# Extract specs for each item group
	for idx, row in enumerate(specs_sheet.iter_rows(values_only=True)):
		if idx == 0:  # Skip header
			continue
		
		if not row or not row[0]:  # Skip empty rows
			continue
		
		item_group = str(row[0]).strip()
		item_group_specs = set()
		
		# Check each spec label column
		for spec_info in spec_labels:
			col_idx = spec_info["index"]
			
			if len(row) > col_idx and row[col_idx] is not None:
				description = str(row[col_idx]).strip()
				if description:  # Only add if not empty
					item_group_specs.add(spec_info["label"])
		
		if item_group_specs:
			specs_by_item_group[item_group] = item_group_specs
	
	frappe.logger().info(f"Extracted specs for {len(specs_by_item_group)} item groups from third sheet")
	return specs_by_item_group


def has_price_in_column(row, price_column_idx):
	"""Check if row has a valid price (> 0) in specified column"""
	if len(row) > price_column_idx and row[price_column_idx] is not None:
		try:
			price_value = float(row[price_column_idx])
			return price_value > 0
		except (ValueError, TypeError):
			return False
	return False


def extract_data_for_price_list(variants_sheet, price_column_idx, column_indices, specs_by_item_group):
	"""
	Extract unique specs, categories, brands, and availability rules
	for items with prices in the specified price list column
	
	Args:
		variants_sheet: Worksheet object
		price_column_idx: Index of price list column
		column_indices: Dict of important column indices
		specs_by_item_group: Dict mapping item groups to their specs
	
	Returns:
		dict: Payload with SpecsExternalIds, CategoriesExternalIds, etc.
	"""
	specs_external_ids = set()
	categories_external_ids = set()
	brands_external_ids = set()
	availability_rules_aliases = set()
	
	for idx, row in enumerate(variants_sheet.iter_rows(values_only=True)):
		if idx == 0:  # Skip header
			continue
		
		# Only process items with prices
		if not has_price_in_column(row, price_column_idx):
			continue
		
		# Extract Item Group (Category)
		if column_indices['item_group'] != -1:
			if len(row) > column_indices['item_group'] and row[column_indices['item_group']]:
				item_group = str(row[column_indices['item_group']]).strip()
				if item_group:
					categories_external_ids.add(item_group)
					
					# Extract specs for this item group
					if item_group in specs_by_item_group:
						specs_external_ids.update(specs_by_item_group[item_group])
		
		# Extract Brand
		if column_indices['brand'] != -1:
			if len(row) > column_indices['brand'] and row[column_indices['brand']]:
				brand = str(row[column_indices['brand']]).strip()
				if brand:
					brands_external_ids.add(brand)
		
		# Extract Availability Rule
		if column_indices['availability_rule'] != -1:
			if len(row) > column_indices['availability_rule'] and row[column_indices['availability_rule']]:
				rule = str(row[column_indices['availability_rule']]).strip()
				if rule:
					availability_rules_aliases.add(rule)
	
	return {
		"SpecsExternalIds": sorted(list(specs_external_ids)),
		"CategoriesExternalIds": sorted(list(categories_external_ids)),
		"BrandsExternalIds": sorted(list(brands_external_ids)),
		"AvailabilityRulesAliases": sorted(list(availability_rules_aliases))
	}


def call_validation_api(config, payload, price_list_name):
	"""
	Make API call to validate the payload
	
	Args:
		config: Validation config with api_url and api_key
		payload: Data to validate
		price_list_name: Name of price list (for logging)
	
	Returns:
		dict: Result with payload, response, and status
	"""
	try:
		headers = {
			"Content-Type": "application/json",
			"Authorization": f"Bearer {config.api_key}"
		}
		
		response = requests.post(
			config.api_url,
			json=payload,
			headers=headers,
			timeout=30
		)
		
		if response.status_code == 200:
			api_response = response.json()
			return {
				"payload": payload,
				"response": api_response,
				"status": "success"
			}
		else:
			error_msg = f"API returned status {response.status_code}: {response.text}"
			frappe.log_error(
				title=f"API Validation Failed for {price_list_name}",
				message=f"Status: {response.status_code}\nResponse: {response.text}"
			)
			return {
				"payload": payload,
				"response": [],
				"status": "error",
				"error": error_msg
			}
	
	except requests.exceptions.Timeout:
		frappe.log_error(
			title=f"API Timeout for {price_list_name}",
			message="Request timed out"
		)
		return {
			"payload": payload,
			"response": [],
			"status": "error",
			"error": "API request timed out after 30 seconds"
		}
	
	except Exception as api_error:
		frappe.log_error(
			title=f"API Error for {price_list_name}",
			message=frappe.get_traceback()
		)
		return {
			"payload": payload,
			"response": [],
			"status": "error",
			"error": str(api_error)
		}


def process_price_list(price_list_name, price_column_idx, variants_sheet, 
                       column_indices, specs_by_item_group, price_list_to_config):
	# Check if validation config exists
	if price_list_name not in price_list_to_config:
		return None
	
	config = price_list_to_config[price_list_name]
	
	# Extract data for items with prices
	payload = extract_data_for_price_list(
		variants_sheet,
		price_column_idx,
		column_indices,
		specs_by_item_group
	)
	
	# Make API call
	result = call_validation_api(config, payload, price_list_name)
	
	return result


@frappe.whitelist()
def extract_and_validate_excel_data(file_url):
	try:
		# Load workbook
		wb = get_workbook_from_file(file_url)
		
		# Validate required sheets
		validate_required_sheets(wb)
		
		# Get Variants sheet and headers
		variants_sheet = wb['Variants']
		variants_headers = get_sheet_headers(variants_sheet)
		
		# Get price list columns
		price_list_columns, cost_column_index = get_price_list_columns(variants_headers)
		
		# Get validation configurations
		price_list_to_config = get_validation_configs()
		
		# Get column indices
		column_indices = get_column_indices(variants_headers)
		
		# Extract specs from third sheet
		specs_by_item_group = extract_specs_from_third_sheet(wb)
		
		# Process each price list
		results = {}
		no_result = []
		for price_list_name, price_column_idx in price_list_columns.items():
			result = process_price_list(
				price_list_name,
				price_column_idx,
				variants_sheet,
				column_indices,
				specs_by_item_group,
				price_list_to_config
			)
   			
			if result:
				results[price_list_name] = result
			else:
				no_result.append(price_list_name)
    
		frappe.msgprint(f"No validation config found for price lists: <b>{', '.join(no_result)}</b>") if no_result else None
		
		wb.close()
		
		# Log summary
		frappe.logger().info(f"Processed {len(results)} price lists for validation")
		
		return results
		
	except Exception as e:
		frappe.log_error(title="Excel Data Extraction Error", message=frappe.get_traceback())
		frappe.throw(f"Error extracting data from Excel file: {str(e)}")
  

# Add this method to item_from_excel.py
@frappe.whitelist()
def validate_excel_structure(file_url):
	"""
	Comprehensive validation of Excel file structure and content
	Returns validation results for display before price list validation
	"""
	try:
		# Get file content
		file_name = frappe.db.get_value("File", {"file_url": file_url})
		if not file_name:
			frappe.throw("File not found")
		
		file_doc = frappe.get_doc("File", file_name)
		file_content = file_doc.get_content()
		
		# Read Excel
		wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
		
		result = {
			"is_valid": True,
			"file_overview": validate_file_overview(wb),
			"sheet_info": None,
			"mandatory_validation": None,
			"variant_relationships": None,
			"item_summary": None,
			"price_list_summary": None
		}
		
		# If file overview fails, stop here
		if not result["file_overview"]["is_valid"]:
			result["is_valid"] = False
			wb.close()
			return result
		
		# Get sheets
		template_sheet = wb[wb.sheetnames[0]]
		variants_sheet = wb[wb.sheetnames[1]]
		specs_sheet = wb[wb.sheetnames[2]] if len(wb.sheetnames) >= 3 else None
		
		# Sheet information
		result["sheet_info"] = validate_sheet_info(template_sheet, variants_sheet)
		
		# If sheet info fails, stop here
		if result["sheet_info"]["template"]["has_errors"] or result["sheet_info"]["variants"]["has_errors"]:
			result["is_valid"] = False
			wb.close()
			return result
		
		# Mandatory field validation (doesn't stop validation)
		result["mandatory_validation"] = validate_mandatory_fields(template_sheet, variants_sheet)
		
		# Variant relationships
		result["variant_relationships"] = validate_variant_relationships(template_sheet, variants_sheet)
		
		# Item summary (groups, brands, specs, availability rules)
		result["item_summary"] = get_item_summary(template_sheet, variants_sheet, specs_sheet)
		
		# Price list summary
		result["price_list_summary"] = get_price_list_summary(variants_sheet)
		
		wb.close()
		return result
		
	except Exception as e:
		frappe.log_error(title="Excel Structure Validation Error", message=frappe.get_traceback())
		frappe.throw(f"Error validating Excel file: {str(e)}")


def validate_file_overview(wb):
	"""Validate basic file structure"""
	sheet_names = wb.sheetnames
	sheet_count = len(sheet_names)
	
	errors = []
	is_valid = True
	
	# Check for required sheets
	if 'Template' not in sheet_names:
		errors.append("File is missing 'Template' sheet - it must be part of the file even if it doesn't contain any rows")
		is_valid = False
	
	if 'Variants' not in sheet_names:
		errors.append("File is missing 'Variants' sheet - it must be part of the file even if it doesn't contain any rows")
		is_valid = False
	
	return {
		"is_valid": is_valid,
		"sheet_count": sheet_count,
		"sheet_names": sheet_names,
		"errors": errors
	}


def validate_sheet_info(template_sheet, variants_sheet):
	"""Validate sheet structure and headers"""
	required_fields = [
		"Item Code", "Item Name", "Item Group", "Brand"
	]
	
	result = {
		"template": validate_single_sheet(template_sheet, required_fields, "Template"),
		"variants": validate_single_sheet(variants_sheet, required_fields, "Variants")
	}
	
	return result


def validate_single_sheet(sheet, required_fields, sheet_name):
	"""Validate a single sheet"""
	if not sheet:
		return {
			"has_errors": True,
			"missing_fields": required_fields,
			"row_count": 0,
			"column_count": 0,
			"headers": []
		}
	
	# Get headers
	headers = []
	for idx, row in enumerate(sheet.iter_rows(values_only=True)):
		if idx == 0:
			headers = [str(h) if h else '' for h in row]
			break
	
	# Check missing required fields
	missing_fields = [field for field in required_fields if field not in headers]
	
	# Count rows (excluding header and empty rows)
	row_count = 0
	for idx, row in enumerate(sheet.iter_rows(values_only=True)):
		if idx == 0:  # Skip header
			continue
		if any(cell is not None for cell in row):  # Non-empty row
			row_count += 1
	
	return {
		"has_errors": len(missing_fields) > 0,
		"missing_fields": missing_fields,
		"row_count": row_count,
		"column_count": len(headers),
		"headers": headers
	}


def validate_mandatory_fields(template_sheet, variants_sheet):
	"""Check for missing values in mandatory fields"""
	required_fields = ["Item Code", "Item Name", "Item Group"]
	errors = []
	
	# Validate template sheet
	errors.extend(validate_mandatory_in_sheet(template_sheet, required_fields, "Template"))
	
	# Validate variants sheet
	errors.extend(validate_mandatory_in_sheet(variants_sheet, required_fields, "Variants"))
	
	return {
		"has_errors": len(errors) > 0,
		"errors": errors
	}


def validate_mandatory_in_sheet(sheet, required_fields, sheet_name):
	"""Validate mandatory fields in a single sheet"""
	if not sheet:
		return []
	
	# Get headers
	headers = []
	for idx, row in enumerate(sheet.iter_rows(values_only=True)):
		if idx == 0:
			headers = [str(h) if h else '' for h in row]
			break
	
	# Get indices of required fields
	field_indices = {}
	for field in required_fields:
		if field in headers:
			field_indices[field] = headers.index(field)
	
	item_code_idx = headers.index("Item Code") if "Item Code" in headers else 0
	
	# Check each row
	errors = []
	error_by_item = {}
	
	for idx, row in enumerate(sheet.iter_rows(values_only=True)):
		if idx == 0:  # Skip header
			continue
		
		if not any(cell is not None for cell in row):  # Skip empty rows
			continue
		
		item_code = str(row[item_code_idx]) if len(row) > item_code_idx and row[item_code_idx] else f"Row {idx + 1}"
		missing_columns = []
		
		# Check each required field
		for field, field_idx in field_indices.items():
			if len(row) <= field_idx or not row[field_idx]:
				missing_columns.append(field)
		
		if missing_columns:
			if item_code not in error_by_item:
				error_by_item[item_code] = {
					"rows": [],
					"columns": set()
				}
			error_by_item[item_code]["rows"].append(idx + 1)
			error_by_item[item_code]["columns"].update(missing_columns)
	
	# Format errors
	for sku, data in error_by_item.items():
		errors.append({
			"sheet": sheet_name,
			"rows": data["rows"],
			"sku": sku,
			"columns": sorted(list(data["columns"]))
		})
	
	return errors


def validate_variant_relationships(template_sheet, variants_sheet):
	"""Validate variant to template relationships"""
	if not template_sheet or not variants_sheet:
		return {
			"template_count": 0,
			"variant_count": 0,
			"missing_templates": []
		}
	
	# Get template item codes
	template_codes = set()
	for idx, row in enumerate(template_sheet.iter_rows(values_only=True)):
		if idx == 0:  # Get header to find Item Code column
			headers = list(row)
			item_code_idx = headers.index("Item Code") if "Item Code" in headers else 0
			continue
		
		if row and len(row) > item_code_idx and row[item_code_idx]:
			template_codes.add(str(row[item_code_idx]).strip())
	
	# Check variants
	variant_count = 0
	missing_templates = []
	
	for idx, row in enumerate(variants_sheet.iter_rows(values_only=True)):
		if idx == 0:  # Get headers
			headers = list(row)
			item_code_idx = headers.index("Item Code") if "Item Code" in headers else 0
			variant_of_idx = headers.index("Variant Of") if "Variant Of" in headers else -1
			continue
		
		if not row or not any(cell is not None for cell in row):
			continue
		
		variant_count += 1
		
		# Check if Variant Of exists
		if variant_of_idx != -1 and len(row) > variant_of_idx and row[variant_of_idx]:
			template_ref = str(row[variant_of_idx]).strip()
			if template_ref not in template_codes:
				variant_code = str(row[item_code_idx]) if len(row) > item_code_idx and row[item_code_idx] else f"Row {idx + 1}"
				missing_templates.append({
					"variant": variant_code,
					"template": template_ref
				})
	
	return {
		"template_count": len(template_codes),
		"variant_count": variant_count,
		"missing_templates": missing_templates
	}


def get_item_summary(template_sheet, variants_sheet, specs_sheet):
	"""Get summary of item groups, brands, specs, and availability rules"""
	summary = {
		"item_groups": {},
		"brands": {},
		"specs": {},
		"availability_rules": {},
		"website_spec_labels": {}
	}
	
	# Process template sheet
	if template_sheet:
		process_sheet_for_summary(template_sheet, summary, specs_sheet)
	
	# Process variants sheet
	if variants_sheet:
		process_sheet_for_summary(variants_sheet, summary, specs_sheet)
	
	# Extract website specification labels from third sheet
	if specs_sheet:
		extract_website_spec_labels(specs_sheet, summary)
	
	return summary


def process_sheet_for_summary(sheet, summary, specs_sheet):
	"""Process a single sheet for summary data"""
	headers = []
	for idx, row in enumerate(sheet.iter_rows(values_only=True)):
		if idx == 0:
			headers = list(row)
			item_group_idx = headers.index("Item Group") if "Item Group" in headers else -1
			brand_idx = headers.index("Brand") if "Brand" in headers else -1
			availability_idx = headers.index("VariantAvailabilityRule") if "VariantAvailabilityRule" in headers else -1
			item_code_idx = headers.index("Item Code") if "Item Code" in headers else 0
			continue
		
		if not row or not any(cell is not None for cell in row):
			continue
		
		# Item Group
		if item_group_idx != -1 and len(row) > item_group_idx and row[item_group_idx]:
			group = str(row[item_group_idx]).strip()
			summary["item_groups"][group] = summary["item_groups"].get(group, 0) + 1
		
		# Brand
		if brand_idx != -1 and len(row) > brand_idx and row[brand_idx]:
			brand = str(row[brand_idx]).strip()
			summary["brands"][brand] = summary["brands"].get(brand, 0) + 1
		
		# Availability Rule
		if availability_idx != -1 and len(row) > availability_idx and row[availability_idx]:
			rule = str(row[availability_idx]).strip()
			summary["availability_rules"][rule] = summary["availability_rules"].get(rule, 0) + 1
		
		# Specs from third sheet
		if specs_sheet and item_group_idx != -1 and len(row) > item_group_idx and row[item_group_idx]:
			item_group = str(row[item_group_idx]).strip()
			specs = get_specs_for_item_group(specs_sheet, item_group)
			for spec in specs:
				summary["specs"][spec] = summary["specs"].get(spec, 0) + 1


def get_specs_for_item_group(specs_sheet, item_group):
	"""Get specification labels for an item group from third sheet"""
	specs = []
	
	# Get headers
	spec_headers = []
	for idx, row in enumerate(specs_sheet.iter_rows(values_only=True)):
		if idx == 0:
			spec_headers = list(row)
			break
	
	if not spec_headers or len(spec_headers) < 2:
		return specs
	
	# Find row for this item group
	for idx, row in enumerate(specs_sheet.iter_rows(values_only=True)):
		if idx == 0:  # Skip header
			continue
		
		if not row or not row[0]:
			continue
		
		if str(row[0]).strip() == item_group:
			# Check which spec columns have values
			for i, header in enumerate(spec_headers[1:], start=1):
				if header and len(row) > i and row[i] is not None:
					val = str(row[i]).strip()
					if val:
						specs.append(str(header).strip())
			break
	
	return specs


def extract_website_spec_labels(specs_sheet, summary):
	"""Extract all website specification labels from third sheet and count usage"""
	if not specs_sheet:
		return
	
	# Get headers (spec labels)
	spec_headers = []
	for idx, row in enumerate(specs_sheet.iter_rows(values_only=True)):
		if idx == 0:
			spec_headers = list(row)
			break
	
	if not spec_headers or len(spec_headers) < 2:
		return
	
	# Extract labels from columns 2+
	spec_label_columns = {}
	for i, header in enumerate(spec_headers[1:], start=1):
		if header and str(header).strip():
			spec_label_columns[i] = str(header).strip()
	
	# Count how many item groups use each label
	label_counts = {}
	
	for idx, row in enumerate(specs_sheet.iter_rows(values_only=True)):
		if idx == 0:  # Skip header
			continue
		
		if not row or not row[0]:  # Skip empty rows
			continue
		
		# For each label column, check if it has a value
		for col_idx, label in spec_label_columns.items():
			if len(row) > col_idx and row[col_idx] is not None:
				val = str(row[col_idx]).strip()
				if val:  # Has a value
					label_counts[label] = label_counts.get(label, 0) + 1
	
	# Store in summary
	summary["website_spec_labels"] = label_counts


def get_price_list_summary(variants_sheet):
	"""Get summary of price list assignments"""
	if not variants_sheet:
		return {"price_lists": {}}
	
	# Get headers
	headers = []
	for idx, row in enumerate(variants_sheet.iter_rows(values_only=True)):
		if idx == 0:
			headers = list(row)
			break
	
	# Find Cost column
	cost_idx = headers.index("Cost") if "Cost" in headers else -1
	if cost_idx == -1:
		return {"price_lists": {}}
	
	# Get price list columns (after Cost)
	price_list_cols = {}
	for i, header in enumerate(headers):
		if i > cost_idx and header and str(header).strip():
			price_list_cols[str(header).strip()] = i
	
	# Count items by price list combination
	price_list_combos = {}
	
	for idx, row in enumerate(variants_sheet.iter_rows(values_only=True)):
		if idx == 0:  # Skip header
			continue
		
		if not row or not any(cell is not None for cell in row):
			continue
		
		# Check which price lists have prices
		item_price_lists = []
		for pl_name, pl_idx in price_list_cols.items():
			if len(row) > pl_idx and row[pl_idx] is not None:
				try:
					price = float(row[pl_idx])
					if price > 0:
						item_price_lists.append(pl_name)
				except (ValueError, TypeError):
					pass
		
		if item_price_lists:
			combo_key = ", ".join(sorted(item_price_lists))
			price_list_combos[combo_key] = price_list_combos.get(combo_key, 0) + 1
	
	return {"price_lists": price_list_combos}